"""
Nexus ERP - FastAPI Main Application
All routes, API endpoints, and business logic
"""
import os, uuid, json, io, urllib.parse
from typing import Optional
from fastapi import FastAPI, Request, Form, Depends, HTTPException, status
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, joinedload

# ReportLab components for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Openpyxl for Excel generation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import func
from dotenv import load_dotenv
import pandas as pd
import numpy as np
import database as db

load_dotenv()

app = FastAPI(title="Nexus ERP", description="The Ultimate AI-Powered Accounting Platform")

# Boot DB safely
try:
    db.init_db()
except Exception as e:
    print(f"Database init warning: {e}")

# Setup static files and templates
try:
    os.makedirs("static/css", exist_ok=True)
    os.makedirs("static/js", exist_ok=True)
    os.makedirs("static/img", exist_ok=True)
    os.makedirs("templates", exist_ok=True)
except Exception:
    pass

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# ──────────────────────────────────────────────────────
# SEO CRAWLER ENDPOINTS (Google, Bing & AI Indexers)
# ──────────────────────────────────────────────────────
@app.get("/sitemap.xml", include_in_schema=False)
async def get_sitemap():
    return FileResponse("static/sitemap.xml", media_type="application/xml")

@app.get("/robots.txt", include_in_schema=False)
async def get_robots():
    return FileResponse("static/robots.txt", media_type="text/plain")



# ──────────────────────────────────────────────────────
# AUTH HELPERS (Session via Cookie - Simple for MVP)
# ──────────────────────────────────────────────────────
class NotAuthenticatedException(Exception):
    """Raised when user is not logged in."""
    pass

@app.exception_handler(NotAuthenticatedException)
async def not_authenticated_handler(request: Request, exc: NotAuthenticatedException):
    return RedirectResponse(url="/auth", status_code=303)

def get_current_user(request: Request) -> Optional[dict]:
    user_cookie = request.cookies.get("nexus_user")
    if not user_cookie:
        return None
    try:
        decoded = urllib.parse.unquote(user_cookie)
        return json.loads(decoded)
    except Exception:
        return None

def require_user(request: Request) -> dict:
    user = get_current_user(request)
    if not user:
        raise NotAuthenticatedException()
    return user


# ──────────────────────────────────────────────────────
# AUTH ROUTES
# ──────────────────────────────────────────────────────
@app.get("/auth", response_class=HTMLResponse)
async def auth_page(request: Request):
    return templates.TemplateResponse(request, "auth.html", {})

@app.post("/auth/register")
async def register(
    request: Request,
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    role: str = Form("CEO"),
    plan: str = Form("corporate"),
    company: str = Form("Nexus Enterprise"),
    industry: str = Form("General Services"),
    tax_id: str = Form(""),
    business_type: str = Form("Private Limited"),
    annual_revenue: str = Form("Rs 1M - 10M"),
    currency: str = Form("USD"),
    language: str = Form("en")
):
    user_id = db.create_user(
        username=username, email=email, password=password, role=role, plan=plan,
        company=company, industry=industry, tax_id=tax_id, business_type=business_type,
        annual_revenue=annual_revenue, currency=currency, language=language
    )
    if not user_id:
        return templates.TemplateResponse(request, "auth.html", {
            "error": "Username or email already exists."
        })
    user_data = db.authenticate_user(username, password)
    target = "/ceo" if role == "CEO" else ("/cfo" if role == "CFO" else "/")
    resp = RedirectResponse(url=target, status_code=303)
    encoded = urllib.parse.quote(json.dumps(user_data))
    resp.set_cookie("nexus_user", encoded, max_age=86400 * 7)
    return resp


@app.post("/auth/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    user_data = db.authenticate_user(username, password)
    if not user_data:
        return templates.TemplateResponse(request, "auth.html", {
            "error": "Invalid username or password. Demo accounts: ceo/password123, cfo/password123, accountant/password123"
        })
    # Redirect based on user role
    role = user_data.get("role", "CEO")
    target_url = "/ceo" if role == "CEO" else ("/cfo" if role == "CFO" else "/")
    resp = RedirectResponse(url=target_url, status_code=303)
    encoded = urllib.parse.quote(json.dumps(user_data))
    resp.set_cookie("nexus_user", encoded, max_age=86400 * 7)
    return resp

@app.post("/auth/switch_role")
async def switch_role(request: Request, new_role: str = Form(...), session: Session = Depends(db.get_db)):
    user = require_user(request)
    if new_role in ["CEO", "CFO", "Accountant"]:
        user["role"] = new_role
        db_user = session.query(db.User).filter(db.User.id == user["id"]).first()
        if db_user:
            db_user.role = new_role
            session.commit()
    target = "/ceo" if new_role == "CEO" else ("/cfo" if new_role == "CFO" else "/")
    resp = RedirectResponse(url=target, status_code=303)
    encoded = urllib.parse.quote(json.dumps(user))
    resp.set_cookie("nexus_user", encoded, max_age=86400 * 7)
    return resp

@app.get("/auth/logout")
async def logout():
    resp = RedirectResponse(url="/auth", status_code=303)
    resp.delete_cookie("nexus_user")
    return resp

@app.get("/auth/login")
async def login_get():
    return RedirectResponse(url="/auth", status_code=303)

@app.get("/auth/register")
async def register_get():
    return RedirectResponse(url="/auth", status_code=303)



# ──────────────────────────────────────────────────────
# DASHBOARD
# ──────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, session: Session = Depends(db.get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/auth", status_code=303)
    
    uid = user["id"]
    
    # KPI Calculations
    accounts = session.query(db.Account).filter(db.Account.user_id == uid).all()
    
    total_assets      = sum(a.balance for a in accounts if a.account_type == "Asset" and a.balance > 0)
    total_liabilities = sum(a.balance for a in accounts if a.account_type == "Liability")
    total_revenue     = sum(a.balance for a in accounts if a.account_type == "Revenue")
    total_expenses    = sum(a.balance for a in accounts if a.account_type == "Expense")
    net_profit        = total_revenue - total_expenses
    equity            = total_assets - total_liabilities
    
    recent_entries = session.query(db.JournalEntry).filter(
        db.JournalEntry.user_id == uid
    ).order_by(db.JournalEntry.created_at.desc()).limit(8).all()
    
    pending_invoices = session.query(db.Invoice).filter(
        db.Invoice.user_id == uid,
        db.Invoice.status.in_(["sent", "overdue"])
    ).count()
    
    unmatched_bank = session.query(db.BankTransaction).filter(
        db.BankTransaction.user_id == uid,
        db.BankTransaction.status == "unmatched"
    ).count()

    operating_cf = round(total_revenue * 0.75 - total_expenses * 0.5, 2)
    investing_cf = -18500.00
    financing_cf = 12000.00
    net_cash_flow = round(operating_cf + investing_cf + financing_cf, 2)
    
    return templates.TemplateResponse(request, "index.html", {
        "user": user,
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "total_revenue": total_revenue,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "equity": equity,
        "operating_cf": operating_cf,
        "investing_cf": investing_cf,
        "financing_cf": financing_cf,
        "net_cash_flow": net_cash_flow,
        "recent_entries": recent_entries,
        "pending_invoices": pending_invoices,
        "unmatched_bank": unmatched_bank,
        "active_page": "dashboard",
    })



# ──────────────────────────────────────────────────────
# CHART OF ACCOUNTS
# ──────────────────────────────────────────────────────
@app.get("/accounts", response_class=HTMLResponse)
async def accounts_page(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    accounts = session.query(db.Account).filter(
        db.Account.user_id == user["id"],
        db.Account.is_active == True
    ).order_by(db.Account.code).all()
    
    return templates.TemplateResponse(request, "accounts.html", {
        "user": user, "accounts": accounts,
        "active_page": "accounts"
    })

@app.post("/api/accounts")
async def create_account(
    request: Request,
    code: str = Form(...),
    name: str = Form(...),
    account_type: str = Form(...),
    sub_type: str = Form(""),
    description: str = Form(""),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    acct = db.Account(
        user_id=user["id"], code=code, name=name,
        account_type=account_type, sub_type=sub_type, description=description
    )
    session.add(acct)
    session.commit()
    return JSONResponse({"success": True, "message": f"Account '{name}' created."})


# ──────────────────────────────────────────────────────
# JOURNAL ENTRIES
# ──────────────────────────────────────────────────────
@app.get("/journal", response_class=HTMLResponse)
async def journal_page(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    entries = session.query(db.JournalEntry).filter(
        db.JournalEntry.user_id == user["id"]
    ).order_by(db.JournalEntry.date.desc()).all()
    
    accounts = session.query(db.Account).filter(
        db.Account.user_id == user["id"],
        db.Account.is_active == True
    ).order_by(db.Account.code).all()

    # Attach lines total for display
    for e in entries:
        e.total = sum(l.debit for l in e.lines)
    
    return templates.TemplateResponse(request, "journal.html", {
        "user": user, "entries": entries,
        "accounts": accounts, "active_page": "journal"
    })

@app.post("/api/journal")
async def post_journal_entry(
    request: Request,
    date: str = Form(...),
    description: str = Form(...),
    reference: str = Form(""),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    body = await request.form()
    
    # Parse multi-line entry: account_id[], debit[], credit[]
    account_ids = body.getlist("account_id[]")
    debits      = body.getlist("debit[]")
    credits     = body.getlist("credit[]")
    
    if not account_ids:
        return JSONResponse({"success": False, "error": "No lines provided."}, status_code=400)
    
    total_debits  = sum(float(d or 0) for d in debits)
    total_credits = sum(float(c or 0) for c in credits)
    
    # Double-entry validation: Debits MUST equal Credits
    if round(total_debits, 2) != round(total_credits, 2):
        return JSONResponse({"success": False, "error": f"Debits ({total_debits:.2f}) must equal Credits ({total_credits:.2f})."}, status_code=400)
    
    entry = db.JournalEntry(
        user_id=user["id"], date=date, description=description, reference=reference
    )
    session.add(entry)
    session.flush()
    
    for i, acct_id in enumerate(account_ids):
        dr = float(debits[i] or 0)
        cr = float(credits[i] or 0)
        if dr == 0 and cr == 0:
            continue
        
        line = db.JournalLine(entry_id=entry.id, account_id=int(acct_id), debit=dr, credit=cr)
        session.add(line)
        
        # Update account balance
        acct = session.query(db.Account).get(int(acct_id))
        if acct:
            # Normal balance rules
            if acct.account_type in ["Asset", "Expense"]:
                acct.balance += dr - cr
            else:  # Liability, Equity, Revenue
                acct.balance += cr - dr
    
    session.commit()
    db._log(session, user["id"], "JOURNAL_ENTRY", f"Entry: {description} | Dr={total_debits:.2f} Cr={total_credits:.2f}")
    
    return JSONResponse({"success": True, "message": "Journal entry recorded.", "entry_id": entry.id, "amount": total_debits})


# ──────────────────────────────────────────────────────
# CONTACTS (Customers & Suppliers)
# ──────────────────────────────────────────────────────
@app.get("/contacts", response_class=HTMLResponse)
async def contacts_page(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    contacts = session.query(db.Contact).filter(
        db.Contact.user_id == user["id"]
    ).order_by(db.Contact.contact_type, db.Contact.name).all()
    
    customers = [c for c in contacts if c.contact_type == "customer"]
    suppliers = [c for c in contacts if c.contact_type == "supplier"]
    
    return templates.TemplateResponse(request, "contacts.html", {
        "user": user, "contacts": contacts,
        "customers": customers, "suppliers": suppliers,
        "active_page": "contacts"
    })

@app.post("/api/contacts")
async def create_contact(
    request: Request,
    contact_type: str = Form(...),
    name: str = Form(...),
    email: str = Form(""),
    phone: str = Form(""),
    address: str = Form(""),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    contact = db.Contact(
        user_id=user["id"], contact_type=contact_type,
        name=name, email=email, phone=phone, address=address
    )
    session.add(contact)
    session.commit()
    return JSONResponse({"success": True, "message": f"{contact_type.title()} '{name}' added.", "id": contact.id})

@app.delete("/api/contacts/{id}")
async def delete_contact(id: int, request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    contact = session.query(db.Contact).filter(db.Contact.id == id, db.Contact.user_id == user["id"]).first()
    if not contact:
        return JSONResponse({"success": False, "error": "Contact not found."}, status_code=404)
    session.delete(contact)
    session.commit()
    return JSONResponse({"success": True, "message": "Contact deleted."})


# ──────────────────────────────────────────────────────
# INVOICES
# ──────────────────────────────────────────────────────
@app.get("/invoices", response_class=HTMLResponse)
async def invoices_page(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    invoices  = session.query(db.Invoice).filter(db.Invoice.user_id == user["id"]).order_by(db.Invoice.date.desc()).all()
    customers = session.query(db.Contact).filter(db.Contact.user_id == user["id"], db.Contact.contact_type == "customer").all()
    
    total_outstanding = sum(i.total - i.amount_paid for i in invoices if i.status in ["sent", "overdue"])
    
    return templates.TemplateResponse(request, "invoices.html", {
        "user": user, "invoices": invoices,
        "customers": customers, "total_outstanding": total_outstanding,
        "active_page": "invoices"
    })

@app.post("/api/invoices")
async def create_invoice(
    request: Request,
    contact_id: int = Form(None),
    customer_name: str = Form(""),
    date: str = Form(...),
    due_date: str = Form(""),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    body = await request.form()
    
    descriptions = body.getlist("item_desc[]")
    quantities   = body.getlist("item_qty[]")
    prices       = body.getlist("item_price[]")
    tax_rates    = body.getlist("item_tax[]")
    
    # Get or create contact
    if not contact_id and customer_name:
        contact = db.Contact(user_id=user["id"], contact_type="customer", name=customer_name)
        session.add(contact)
        session.flush()
        contact_id = contact.id
    
    subtotal = 0.0
    tax_total = 0.0
    
    # Count existing invoices to number sequentially
    count = session.query(db.Invoice).filter(db.Invoice.user_id == user["id"]).count()
    inv_no = f"INV-{count + 1:04d}"
    
    invoice = db.Invoice(
        user_id=user["id"], contact_id=contact_id, invoice_no=inv_no,
        date=date, due_date=due_date, status="sent"
    )
    session.add(invoice)
    session.flush()
    
    for i, desc in enumerate(descriptions):
        qty   = float(quantities[i] or 1)
        price = float(prices[i] or 0)
        tax   = float(tax_rates[i] or 0)
        amount = round(qty * price, 2)
        tax_amount = round(amount * tax / 100, 2)
        
        il = db.InvoiceLine(invoice_id=invoice.id, description=desc, quantity=qty,
                            unit_price=price, tax_rate=tax, amount=amount)
        session.add(il)
        subtotal  += amount
        tax_total += tax_amount
    
    invoice.subtotal   = round(subtotal, 2)
    invoice.tax_amount = round(tax_total, 2)
    invoice.total      = round(subtotal + tax_total, 2)
    
    # Auto-post journal entry: Dr Accounts Receivable / Cr Sales Revenue
    ar_account  = session.query(db.Account).filter(db.Account.user_id == user["id"], db.Account.code == "1200").first()
    rev_account = session.query(db.Account).filter(db.Account.user_id == user["id"], db.Account.code == "4000").first()
    
    if ar_account and rev_account:
        je = db.JournalEntry(user_id=user["id"], date=date, description=f"Invoice {inv_no}",
                             reference=inv_no, entry_type="invoice")
        session.add(je)
        session.flush()
        session.add(db.JournalLine(entry_id=je.id, account_id=ar_account.id,
                                   debit=invoice.total, description=f"Invoice {inv_no}"))
        session.add(db.JournalLine(entry_id=je.id, account_id=rev_account.id,
                                   credit=invoice.total, description=f"Sales - Invoice {inv_no}"))
        ar_account.balance  += invoice.total
        rev_account.balance += invoice.total
    
    session.commit()
    return JSONResponse({"success": True, "message": f"Invoice {inv_no} created!", "invoice_no": inv_no, "total": invoice.total})

@app.post("/api/invoices/{id}/pay")
async def pay_invoice(
    id: int,
    request: Request,
    amount: float = Form(...),
    date: str = Form(...),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    invoice = session.query(db.Invoice).filter(db.Invoice.id == id, db.Invoice.user_id == user["id"]).first()
    if not invoice:
        return JSONResponse({"success": False, "error": "Invoice not found."}, status_code=404)
        
    if invoice.status == "paid":
        return JSONResponse({"success": False, "error": "Invoice is already paid."}, status_code=400)
        
    remaining = invoice.total - invoice.amount_paid
    if amount <= 0 or amount > remaining + 0.01:
        return JSONResponse({"success": False, "error": f"Invalid payment amount. Outstanding balance is ${remaining:.2f}."}, status_code=400)
        
    # Update payment details
    invoice.amount_paid = round(invoice.amount_paid + amount, 2)
    if abs(invoice.amount_paid - invoice.total) < 0.01:
        invoice.status = "paid"
    else:
        invoice.status = "sent"  # Keep sent for partial payment
        
    # Post double-entry journal: Dr Cash at Bank (code 1000) / Cr Accounts Receivable (code 1200)
    cash_account = session.query(db.Account).filter(db.Account.user_id == user["id"], db.Account.code == "1000").first()
    ar_account   = session.query(db.Account).filter(db.Account.user_id == user["id"], db.Account.code == "1200").first()
    
    if cash_account and ar_account:
        je = db.JournalEntry(
            user_id=user["id"], date=date,
            description=f"Payment for Invoice {invoice.invoice_no}",
            reference=invoice.invoice_no, entry_type="payment"
        )
        session.add(je)
        session.flush()
        
        session.add(db.JournalLine(entry_id=je.id, account_id=cash_account.id, debit=amount, description=f"Payment - Inv {invoice.invoice_no}"))
        session.add(db.JournalLine(entry_id=je.id, account_id=ar_account.id, credit=amount, description=f"Payment - Inv {invoice.invoice_no}"))
        
        cash_account.balance += amount
        ar_account.balance   -= amount
        
    session.commit()
    db._log(session, user["id"], "PAYMENT_RECORDED", f"Invoice {invoice.invoice_no} paid ${amount:.2f}")
    return JSONResponse({"success": True, "message": f"Payment of ${amount:.2f} recorded successfully."})

@app.get("/api/invoices/{id}/pdf")
async def get_invoice_pdf(id: int, request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    # Fetch invoice
    invoice = session.query(db.Invoice).filter(db.Invoice.id == id, db.Invoice.user_id == user["id"]).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
        
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()
    
    # Create the PDF document
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Custom Styles for premium look
    title_style = ParagraphStyle(
        'InvoiceTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#1e3a8a') # dark blue
    )
    normal_bold = ParagraphStyle(
        'NormalBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=14
    )
    right_align = ParagraphStyle(
        'RightAlign',
        parent=styles['Normal'],
        alignment=2, # Right align
        fontSize=10,
        leading=14
    )
    right_align_bold = ParagraphStyle(
        'RightAlignBold',
        parent=styles['Normal'],
        alignment=2, # Right align
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=14
    )
    
    # Invoice Header Section
    story.append(Paragraph(f"INVOICE", title_style))
    story.append(Spacer(1, 10))
    
    # Company & Invoice Info Table
    info_data = [
        [
            Paragraph(f"<b>Seller:</b><br/>{invoice.user.company_name if invoice.user else 'My Company'}<br/>{invoice.user.email if invoice.user else ''}", styles['Normal']),
            Paragraph(f"<b>Invoice #:</b> {invoice.invoice_no}<br/><b>Date:</b> {invoice.date}<br/><b>Due Date:</b> {invoice.due_date or '—'}<br/><b>Status:</b> {invoice.status.upper()}", right_align)
        ]
    ]
    info_table = Table(info_data, colWidths=[260, 260])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 15))
    
    # Customer Info Section
    cust_data = [
        [
            Paragraph(f"<b>Bill To:</b><br/>{invoice.contact.name if invoice.contact else 'Valued Customer'}<br/>{invoice.contact.email if invoice.contact and invoice.contact.email else ''}<br/>{invoice.contact.address if invoice.contact and invoice.contact.address else ''}", styles['Normal']),
            ""
        ]
    ]
    cust_table = Table(cust_data, colWidths=[260, 260])
    cust_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 15),
    ]))
    story.append(cust_table)
    
    # Line items Table
    items_data = [
        [
            Paragraph("<b>Description</b>", normal_bold),
            Paragraph("<b>Qty</b>", normal_bold),
            Paragraph("<b>Unit Price</b>", right_align_bold),
            Paragraph("<b>Amount</b>", right_align_bold)
        ]
    ]
    
    for line in invoice.lines:
        items_data.append([
            Paragraph(line.description, styles['Normal']),
            Paragraph(str(int(line.quantity)), styles['Normal']),
            Paragraph(f"${line.unit_price:,.2f}", right_align),
            Paragraph(f"${line.amount:,.2f}", right_align)
        ])
        
    # Totals rows
    items_data.append(["", "", Paragraph("<b>Subtotal:</b>", right_align), Paragraph(f"${invoice.subtotal:,.2f}", right_align)])
    items_data.append(["", "", Paragraph("<b>Tax:</b>", right_align), Paragraph(f"${invoice.tax_amount:,.2f}", right_align)])
    items_data.append(["", "", Paragraph("<b>Total:</b>", right_align_bold), Paragraph(f"<b>${invoice.total:,.2f}</b>", right_align_bold)])
    items_data.append(["", "", Paragraph("<b>Amount Paid:</b>", right_align), Paragraph(f"${invoice.amount_paid:,.2f}", right_align)])
    items_data.append(["", "", Paragraph("<b>Balance Due:</b>", right_align_bold), Paragraph(f"<b>${(invoice.total - invoice.amount_paid):,.2f}</b>", right_align_bold)])
    
    items_table = Table(items_data, colWidths=[280, 50, 90, 100])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('ALIGN', (0,0), (-1,0), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor('#cbd5e1')),
        ('LINEBELOW', (0,1), (-1, len(invoice.lines)), 0.5, colors.HexColor('#e2e8f0')),
        ('LINEBELOW', (2, -5), (3, -1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    
    story.append(items_table)
    
    # Build document
    doc.build(story)
    
    # Grab the byte stream and reset pointer
    buffer.seek(0)
    
    # Return PDF response
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Invoice_{invoice.invoice_no}.pdf"}
    )


# ──────────────────────────────────────────────────────
# BANK RECONCILIATION
# ──────────────────────────────────────────────────────
@app.get("/reconciliation", response_class=HTMLResponse)
async def reconciliation_page(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    bank_txns = session.query(db.BankTransaction).filter(
        db.BankTransaction.user_id == user["id"]
    ).order_by(db.BankTransaction.date.desc()).limit(50).all()
    
    matched   = sum(1 for t in bank_txns if t.status == "matched")
    suggested = sum(1 for t in bank_txns if t.status == "suggested")
    unmatched = sum(1 for t in bank_txns if t.status == "unmatched")
    
    return templates.TemplateResponse(request, "reconciliation.html", {
        "user": user, "transactions": bank_txns,
        "matched": matched, "suggested": suggested, "unmatched": unmatched,
        "active_page": "reconciliation"
    })

@app.post("/api/bank/import")
async def import_bank(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    form = await request.form()
    file = form.get("file")
    
    if not file:
        return JSONResponse({"success": False, "error": "No file uploaded."}, status_code=400)
    
    contents = await file.read()
    try:
        import io
        if file.filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        return JSONResponse({"success": False, "error": f"Could not read file: {e}"}, status_code=400)
    
    df.columns = [c.strip().lower() for c in df.columns]
    
    if "date" not in df.columns or "amount" not in df.columns:
        return JSONResponse({"success": False, "error": "File must contain 'date' and 'amount' columns."}, status_code=400)
    
    if "description" not in df.columns:
        df["description"] = "Imported Transaction"
    if "reference" not in df.columns:
        df["reference"] = ""
    
    df = df.dropna(subset=["date", "amount"])
    count = 0
    for _, row in df.iterrows():
        txn = db.BankTransaction(
            user_id=user["id"],
            date=str(row["date"]),
            description=str(row.get("description", "")),
            reference=str(row.get("reference", "")),
            amount=float(row["amount"]),
            balance=float(row["balance"]) if "balance" in row and pd.notna(row["balance"]) else None
        )
        session.add(txn)
        count += 1
    
    session.commit()
    return JSONResponse({"success": True, "message": f"{count} transactions imported."})

@app.post("/api/bank/reconcile")
async def auto_reconcile(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    uid = user["id"]
    
    bank_txns = session.query(db.BankTransaction).filter(
        db.BankTransaction.user_id == uid,
        db.BankTransaction.status == "unmatched"
    ).all()
    
    je_lines = session.query(db.JournalLine).join(db.JournalEntry).filter(
        db.JournalEntry.user_id == uid
    ).all()
    
    matched_count = 0
    matched_bank_ids = set()
    
    for txn in bank_txns:
        if txn.id in matched_bank_ids:
            continue
        
        for jl in je_lines:
            entry_amount = jl.debit if txn.amount > 0 else jl.credit
            if abs(entry_amount - abs(txn.amount)) < 0.01:
                txn.status   = "matched"
                txn.match_id = str(jl.entry_id)
                matched_bank_ids.add(txn.id)
                matched_count += 1
                break
    
    session.commit()
    return JSONResponse({"success": True, "matched": matched_count,
                         "message": f"Auto-reconciled {matched_count} transactions."})


# ──────────────────────────────────────────────────────
# FINANCIAL REPORTS
# ──────────────────────────────────────────────────────
@app.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    uid = user["id"]
    data = _get_report_data(uid, session)
    
    return templates.TemplateResponse(request, "reports.html", {
        "user": user,
        "accounts":      data["accounts"],
        "revenues":      data["revenues"],
        "expenses":      data["expenses"],
        "total_rev":     data["total_rev"],
        "total_exp":     data["total_exp"],
        "net_profit":    data["net_profit"],
        "assets":        data["assets"],
        "liabilities":   data["liabilities"],
        "equity_accts":  data["equity_accts"],
        "total_assets":  data["total_assets"],
        "total_liab":    data["total_liab"],
        "total_equity":  data["total_equity"],
        "total_debits":  data["total_debits"],
        "total_credits": data["total_credits"],
        "active_page": "reports"
    })

def _get_report_data(uid: int, session: Session):
    accounts = session.query(db.Account).filter(db.Account.user_id == uid).order_by(db.Account.code).all()
    # Trial Balance
    total_debits   = sum(max(a.balance, 0) for a in accounts if a.account_type in ["Asset", "Expense"])
    total_credits  = sum(max(a.balance, 0) for a in accounts if a.account_type in ["Liability", "Equity", "Revenue"])
    
    # P&L
    revenues  = [a for a in accounts if a.account_type == "Revenue"]
    expenses  = [a for a in accounts if a.account_type == "Expense"]
    total_rev = sum(a.balance for a in revenues)
    total_exp = sum(a.balance for a in expenses)
    net_profit = total_rev - total_exp
    
    # Balance Sheet
    assets      = [a for a in accounts if a.account_type == "Asset"]
    liabilities = [a for a in accounts if a.account_type == "Liability"]
    equity_accts = [a for a in accounts if a.account_type == "Equity"]
    
    total_assets = sum(a.balance for a in assets)
    total_liab   = sum(a.balance for a in liabilities)
    total_equity = sum(a.balance for a in equity_accts) + net_profit  # Retained earnings
    
    return {
        "accounts": accounts,
        "revenues": revenues, "expenses": expenses,
        "total_rev": total_rev, "total_exp": total_exp, "net_profit": net_profit,
        "assets": assets, "liabilities": liabilities, "equity_accts": equity_accts,
        "total_assets": total_assets, "total_liab": total_liab, "total_equity": total_equity,
        "total_debits": total_debits, "total_credits": total_credits
    }

@app.get("/api/reports/pdf")
async def get_report_pdf(type: str = "pl", request: Request = None, session: Session = Depends(db.get_db)):
    user = require_user(request)
    uid = user["id"]
    data = _get_report_data(uid, session)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=20, leading=24,
        textColor=colors.HexColor('#1e3a8a'), alignment=1
    )
    subtitle_style = ParagraphStyle(
        'ReportSubtitle', parent=styles['Normal'],
        fontSize=10, leading=14, textColor=colors.HexColor('#475569'), alignment=1
    )
    bold_style = ParagraphStyle('BoldStyle', parent=styles['Normal'], fontName='Helvetica-Bold')
    right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], alignment=2)
    right_bold_style = ParagraphStyle('RightBoldStyle', parent=styles['Normal'], fontName='Helvetica-Bold', alignment=2)
    
    company_name = user.get("company", "My Company")
    
    if type == "pl":
        story.append(Paragraph(f"{company_name}", title_style))
        story.append(Paragraph("PROFIT & LOSS STATEMENT", subtitle_style))
        story.append(Paragraph("All Periods", subtitle_style))
        story.append(Spacer(1, 20))
        
        table_data = [[Paragraph("<b>Account</b>", bold_style), Paragraph("<b>Type</b>", bold_style), Paragraph("<b>Amount</b>", right_bold_style)]]
        
        # Revenues
        table_data.append([Paragraph("<b>REVENUE</b>", bold_style), "", ""])
        for a in data["revenues"]:
            table_data.append([Paragraph(f"{a.code} — {a.name}", styles['Normal']), Paragraph(a.account_type, styles['Normal']), Paragraph(f"${a.balance:,.2f}", right_style)])
        table_data.append(["", Paragraph("<b>Total Revenue</b>", bold_style), Paragraph(f"<b>${data['total_rev']:,.2f}</b>", right_bold_style)])
        
        # Expenses
        table_data.append([Paragraph("<b>EXPENSES</b>", bold_style), "", ""])
        for a in data["expenses"]:
            table_data.append([Paragraph(f"{a.code} — {a.name}", styles['Normal']), Paragraph(a.account_type, styles['Normal']), Paragraph(f"${a.balance:,.2f}", right_style)])
        table_data.append(["", Paragraph("<b>Total Expenses</b>", bold_style), Paragraph(f"<b>${data['total_exp']:,.2f}</b>", right_bold_style)])
        
        # Net Profit
        label = "NET PROFIT" if data["net_profit"] >= 0 else "NET LOSS"
        table_data.append([Paragraph(f"<b>{label}</b>", bold_style), "", Paragraph(f"<b>${abs(data['net_profit']):,.2f}</b>", right_bold_style)])
        
        t = Table(table_data, colWidths=[280, 100, 140])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
            ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor('#cbd5e1')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('LINEBELOW', (0, -1), (-1, -1), 1.5, colors.HexColor('#10b981' if data['net_profit'] >= 0 else '#ef4444')),
        ]))
        story.append(t)
        
    elif type == "bs":
        story.append(Paragraph(f"{company_name}", title_style))
        story.append(Paragraph("BALANCE SHEET", subtitle_style))
        story.append(Paragraph("As of Today", subtitle_style))
        story.append(Spacer(1, 20))
        
        table_data = [[Paragraph("<b>Account</b>", bold_style), Paragraph("<b>Type</b>", bold_style), Paragraph("<b>Amount</b>", right_bold_style)]]
        
        # Assets
        table_data.append([Paragraph("<b>ASSETS</b>", bold_style), "", ""])
        for a in data["assets"]:
            table_data.append([Paragraph(f"{a.code} — {a.name}", styles['Normal']), Paragraph(a.sub_type or a.account_type, styles['Normal']), Paragraph(f"${a.balance:,.2f}", right_style)])
        table_data.append(["", Paragraph("<b>Total Assets</b>", bold_style), Paragraph(f"<b>${data['total_assets']:,.2f}</b>", right_bold_style)])
        
        # Liabilities
        table_data.append([Paragraph("<b>LIABILITIES</b>", bold_style), "", ""])
        for a in data["liabilities"]:
            table_data.append([Paragraph(f"{a.code} — {a.name}", styles['Normal']), Paragraph(a.sub_type or a.account_type, styles['Normal']), Paragraph(f"${a.balance:,.2f}", right_style)])
        table_data.append(["", Paragraph("<b>Total Liabilities</b>", bold_style), Paragraph(f"<b>${data['total_liab']:,.2f}</b>", right_bold_style)])
        
        # Equity
        table_data.append([Paragraph("<b>EQUITY</b>", bold_style), "", ""])
        for a in data["equity_accts"]:
            table_data.append([Paragraph(f"{a.code} — {a.name}", styles['Normal']), Paragraph(a.sub_type or a.account_type, styles['Normal']), Paragraph(f"${a.balance:,.2f}", right_style)])
        # Current earnings
        table_data.append([Paragraph("Current Year Earnings (Net Profit)", styles['Normal']), Paragraph("Retained Earnings", styles['Normal']), Paragraph(f"${data['net_profit']:,.2f}", right_style)])
        table_data.append(["", Paragraph("<b>Total Equity</b>", bold_style), Paragraph(f"<b>${data['total_equity']:,.2f}</b>", right_bold_style)])
        
        # Balance check
        balanced = "BALANCED" if round(data["total_assets"], 2) == round(data["total_liab"] + data["total_equity"], 2) else "UNBALANCED"
        table_data.append([Paragraph(f"<b>Equation Check ({balanced})</b>", bold_style), Paragraph(f"Liabilities + Equity: ${data['total_liab'] + data['total_equity']:,.2f}", styles['Normal']), Paragraph(f"Assets: ${data['total_assets']:,.2f}", right_bold_style)])
        
        t = Table(table_data, colWidths=[280, 120, 120])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
            ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor('#cbd5e1')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f8fafc')),
        ]))
        story.append(t)
        
    elif type == "tb":
        story.append(Paragraph(f"{company_name}", title_style))
        story.append(Paragraph("TRIAL BALANCE", subtitle_style))
        story.append(Paragraph("All Periods", subtitle_style))
        story.append(Spacer(1, 20))
        
        table_data = [[
            Paragraph("<b>Code</b>", bold_style),
            Paragraph("<b>Account Name</b>", bold_style),
            Paragraph("<b>Type</b>", bold_style),
            Paragraph("<b>Debit ($)</b>", right_bold_style),
            Paragraph("<b>Credit ($)</b>", right_bold_style)
        ]]
        
        for a in data["accounts"]:
            if a.balance != 0:
                dr = f"${a.balance:,.2f}" if a.account_type in ['Asset','Expense'] and a.balance > 0 else "—"
                cr = f"${a.balance:,.2f}" if a.account_type in ['Liability','Equity','Revenue'] and a.balance > 0 else "—"
                table_data.append([
                    Paragraph(a.code, styles['Normal']),
                    Paragraph(a.name, styles['Normal']),
                    Paragraph(a.account_type, styles['Normal']),
                    Paragraph(dr, right_style),
                    Paragraph(cr, right_style)
                ])
                
        table_data.append([
            "", Paragraph("<b>TOTALS</b>", bold_style), "",
            Paragraph(f"<b>${data['total_debits']:,.2f}</b>", right_bold_style),
            Paragraph(f"<b>${data['total_credits']:,.2f}</b>", right_bold_style)
        ])
        
        t = Table(table_data, colWidths=[60, 200, 80, 90, 90])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
            ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor('#cbd5e1')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('LINEBELOW', (0,-1), (-1,-1), 1.5, colors.HexColor('#cbd5e1')),
        ]))
        story.append(t)
        
    doc.build(story)
    buffer.seek(0)
    
    filename = f"Report_{type.upper()}_{company_name.replace(' ', '_')}.pdf"
    return StreamingResponse(
        buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/reports/excel")
async def get_report_excel(type: str = "pl", request: Request = None, session: Session = Depends(db.get_db)):
    user = require_user(request)
    uid = user["id"]
    data = _get_report_data(uid, session)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    company_name = user.get("company", "My Company")
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="475569")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    section_font = Font(name="Calibri", size=11, bold=True)
    total_font = Font(name="Calibri", size=11, bold=True)
    double_bottom_border = Border(bottom=Side(style='double', color="000000"), top=Side(style='thin', color="CCCCCC"))
    thin_border = Border(bottom=Side(style='thin', color="E2E8F0"))
    
    ws.views.sheetView[0].showGridLines = True
    
    if type == "pl":
        ws.title = "Profit & Loss"
        ws.append([company_name])
        ws.append(["Profit & Loss Statement"])
        ws.append(["All Periods"])
        ws.append([])
        
        ws.cell(1,1).font = title_font
        ws.cell(2,1).font = subtitle_font
        ws.cell(3,1).font = subtitle_font
        
        ws.append(["Account Code", "Account Name", "Type", "Balance"])
        for col_idx in range(1, 5):
            cell = ws.cell(5, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            
        ws.append(["REVENUE"])
        ws.cell(ws.max_row, 1).font = section_font
        
        for a in data["revenues"]:
            ws.append([a.code, a.name, a.account_type, a.balance])
            ws.cell(ws.max_row, 4).number_format = "$#,##0.00"
            
        ws.append(["", "Total Revenue", "", data["total_rev"]])
        row_idx = ws.max_row
        ws.cell(row_idx, 2).font = total_font
        ws.cell(row_idx, 4).font = total_font
        ws.cell(row_idx, 4).number_format = "$#,##0.00"
        ws.cell(row_idx, 4).border = thin_border
        
        ws.append([])
        ws.append(["EXPENSES"])
        ws.cell(ws.max_row, 1).font = section_font
        
        for a in data["expenses"]:
            ws.append([a.code, a.name, a.account_type, a.balance])
            ws.cell(ws.max_row, 4).number_format = "$#,##0.00"
            
        ws.append(["", "Total Expenses", "", data["total_exp"]])
        row_idx = ws.max_row
        ws.cell(row_idx, 2).font = total_font
        ws.cell(row_idx, 4).font = total_font
        ws.cell(row_idx, 4).number_format = "$#,##0.00"
        ws.cell(row_idx, 4).border = thin_border
        
        ws.append([])
        label = "Net Profit" if data["net_profit"] >= 0 else "Net Loss"
        ws.append(["", label, "", data["net_profit"]])
        row_idx = ws.max_row
        ws.cell(row_idx, 2).font = title_font
        ws.cell(row_idx, 4).font = title_font
        ws.cell(row_idx, 4).number_format = "$#,##0.00"
        ws.cell(row_idx, 4).border = double_bottom_border
        
    elif type == "bs":
        ws.title = "Balance Sheet"
        ws.append([company_name])
        ws.append(["Balance Sheet"])
        ws.append(["As of Today"])
        ws.append([])
        
        ws.cell(1,1).font = title_font
        ws.cell(2,1).font = subtitle_font
        ws.cell(3,1).font = subtitle_font
        
        ws.append(["Account Code", "Account Name", "Sub Type / Type", "Balance"])
        for col_idx in range(1, 5):
            cell = ws.cell(5, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            
        ws.append(["ASSETS"])
        ws.cell(ws.max_row, 1).font = section_font
        
        for a in data["assets"]:
            ws.append([a.code, a.name, a.sub_type or a.account_type, a.balance])
            ws.cell(ws.max_row, 4).number_format = "$#,##0.00"
            
        ws.append(["", "Total Assets", "", data["total_assets"]])
        row_idx = ws.max_row
        ws.cell(row_idx, 2).font = total_font
        ws.cell(row_idx, 4).font = total_font
        ws.cell(row_idx, 4).number_format = "$#,##0.00"
        ws.cell(row_idx, 4).border = thin_border
        
        ws.append([])
        ws.append(["LIABILITIES"])
        ws.cell(ws.max_row, 1).font = section_font
        
        for a in data["liabilities"]:
            ws.append([a.code, a.name, a.sub_type or a.account_type, a.balance])
            ws.cell(ws.max_row, 4).number_format = "$#,##0.00"
            
        ws.append(["", "Total Liabilities", "", data["total_liab"]])
        row_idx = ws.max_row
        ws.cell(row_idx, 2).font = total_font
        ws.cell(row_idx, 4).font = total_font
        ws.cell(row_idx, 4).number_format = "$#,##0.00"
        ws.cell(row_idx, 4).border = thin_border
        
        ws.append([])
        ws.append(["EQUITY"])
        ws.cell(ws.max_row, 1).font = section_font
        
        for a in data["equity_accts"]:
            ws.append([a.code, a.name, a.sub_type or a.account_type, a.balance])
            ws.cell(ws.max_row, 4).number_format = "$#,##0.00"
            
        ws.append(["", "Current Year Earnings (Net Profit)", "Retained Earnings", data["net_profit"]])
        ws.cell(ws.max_row, 4).number_format = "$#,##0.00"
        
        ws.append(["", "Total Equity", "", data["total_equity"]])
        row_idx = ws.max_row
        ws.cell(row_idx, 2).font = total_font
        ws.cell(row_idx, 4).font = total_font
        ws.cell(row_idx, 4).number_format = "$#,##0.00"
        ws.cell(row_idx, 4).border = thin_border
        
        ws.append([])
        balanced = "BALANCED" if round(data["total_assets"], 2) == round(data["total_liab"] + data["total_equity"], 2) else "UNBALANCED"
        ws.append(["Equation Check:", balanced, f"Assets = Liab + Equity", data["total_assets"]])
        row_idx = ws.max_row
        ws.cell(row_idx, 1).font = section_font
        ws.cell(row_idx, 2).font = section_font
        ws.cell(row_idx, 4).font = section_font
        ws.cell(row_idx, 4).number_format = "$#,##0.00"
        ws.cell(row_idx, 4).border = double_bottom_border
        
    elif type == "tb":
        ws.title = "Trial Balance"
        ws.append([company_name])
        ws.append(["Trial Balance"])
        ws.append(["All Periods"])
        ws.append([])
        
        ws.cell(1,1).font = title_font
        ws.cell(2,1).font = subtitle_font
        ws.cell(3,1).font = subtitle_font
        
        ws.append(["Code", "Account Name", "Account Type", "Debit", "Credit"])
        for col_idx in range(1, 6):
            cell = ws.cell(5, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            
        for a in data["accounts"]:
            if a.balance != 0:
                dr = a.balance if a.account_type in ['Asset','Expense'] and a.balance > 0 else 0
                cr = a.balance if a.account_type in ['Liability','Equity','Revenue'] and a.balance > 0 else 0
                
                ws.append([
                    a.code,
                    a.name,
                    a.account_type,
                    dr if dr > 0 else "",
                    cr if cr > 0 else ""
                ])
                if dr > 0: ws.cell(ws.max_row, 4).number_format = "$#,##0.00"
                if cr > 0: ws.cell(ws.max_row, 5).number_format = "$#,##0.00"
                
        ws.append(["", "TOTALS", "", data["total_debits"], data["total_credits"]])
        row_idx = ws.max_row
        ws.cell(row_idx, 2).font = total_font
        ws.cell(row_idx, 4).font = total_font
        ws.cell(row_idx, 5).font = total_font
        ws.cell(row_idx, 4).number_format = "$#,##0.00"
        ws.cell(row_idx, 5).number_format = "$#,##0.00"
        ws.cell(row_idx, 4).border = double_bottom_border
        ws.cell(row_idx, 5).border = double_bottom_border
        
    # Auto fit column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Report_{type.upper()}_{company_name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ──────────────────────────────────────────────────────
# PAYROLL
# ──────────────────────────────────────────────────────
@app.get("/payroll", response_class=HTMLResponse)
async def payroll_page(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    employees = session.query(db.Employee).filter(
        db.Employee.user_id == user["id"],
        db.Employee.is_active == True
    ).all()
    payrolls = session.query(db.Payroll).filter(db.Payroll.user_id == user["id"]).order_by(db.Payroll.period.desc()).limit(20).all()
    
    return templates.TemplateResponse(request, "payroll.html", {
        "user": user,
        "employees": employees, "payrolls": payrolls,
        "active_page": "payroll"
    })

@app.post("/api/payroll/employee")
async def add_employee(
    request: Request,
    name: str = Form(...),
    designation: str = Form(""),
    salary: float = Form(...),
    tax_rate: float = Form(0.0),
    joined_date: str = Form(""),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    emp = db.Employee(user_id=user["id"], name=name, designation=designation,
                      salary=salary, tax_rate=tax_rate, joined_date=joined_date)
    session.add(emp)
    session.commit()
    return JSONResponse({"success": True, "message": f"Employee '{name}' added."})

@app.post("/api/payroll/run")
async def run_payroll(
    request: Request,
    period: str = Form(...),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    uid = user["id"]
    employees = session.query(db.Employee).filter(db.Employee.user_id == uid, db.Employee.is_active == True).all()
    
    salary_account = session.query(db.Account).filter(db.Account.user_id == uid, db.Account.code == "5100").first()
    bank_account   = session.query(db.Account).filter(db.Account.user_id == uid, db.Account.code == "1000").first()
    
    count = 0
    for emp in employees:
        tax = round(emp.salary * emp.tax_rate / 100, 2)
        net = round(emp.salary - tax, 2)
        
        payroll = db.Payroll(
            user_id=uid, employee_id=emp.id, period=period,
            gross_salary=emp.salary, tax_withheld=tax, net_salary=net, paid=True, paid_on=period
        )
        session.add(payroll)
        
        # Journal Entry: Dr Salary Expense / Cr Bank
        if salary_account and bank_account:
            je = db.JournalEntry(user_id=uid, date=period + "-01", description=f"Payroll {period} - {emp.name}",
                                 reference=f"PAYROLL-{period}", entry_type="payroll")
            session.add(je)
            session.flush()
            session.add(db.JournalLine(entry_id=je.id, account_id=salary_account.id, debit=emp.salary))
            session.add(db.JournalLine(entry_id=je.id, account_id=bank_account.id, credit=emp.salary))
            salary_account.balance += emp.salary
            bank_account.balance   -= emp.salary
        
        count += 1
    
    session.commit()
    return JSONResponse({"success": True, "message": f"Payroll run for {count} employees for period {period}."})


# ──────────────────────────────────────────────────────
# FIXED ASSETS & DEPRECIATION
# ──────────────────────────────────────────────────────
@app.get("/fixed_assets", response_class=HTMLResponse)
async def fixed_assets_page(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    assets = session.query(db.FixedAsset).filter(
        db.FixedAsset.user_id == user["id"]
    ).all()
    
    total_cost = sum(a.cost for a in assets)
    total_depr = sum(a.accumulated_depreciation for a in assets)
    net_book_value = total_cost - total_depr
    
    return templates.TemplateResponse(request, "fixed_assets.html", {
        "user": user,
        "assets": assets,
        "total_cost": total_cost,
        "total_depr": total_depr,
        "net_book_value": net_book_value,
        "active_page": "fixed_assets"
    })

@app.post("/api/fixed_assets")
async def create_fixed_asset(
    request: Request,
    name: str = Form(...),
    cost: float = Form(...),
    salvage_value: float = Form(0.0),
    useful_life: int = Form(...),
    purchase_date: str = Form(...),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    asset = db.FixedAsset(
        user_id=user["id"], name=name, cost=cost,
        salvage_value=salvage_value, useful_life=useful_life,
        purchase_date=purchase_date
    )
    session.add(asset)
    session.commit()
    return JSONResponse({"success": True, "message": f"Asset '{name}' registered successfully."})

@app.post("/api/fixed_assets/depreciate")
async def depreciate_assets(
    request: Request,
    period: str = Form(...),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    uid = user["id"]
    
    assets = session.query(db.FixedAsset).filter(db.FixedAsset.user_id == uid).all()
    if not assets:
        return JSONResponse({"success": False, "error": "No assets found to depreciate."}, status_code=400)
    
    depr_expense_account = session.query(db.Account).filter(db.Account.user_id == uid, db.Account.code == "5600").first()
    accum_depr_account   = session.query(db.Account).filter(db.Account.user_id == uid, db.Account.code == "1510").first()
    
    total_depr_amount = 0.0
    
    for asset in assets:
        monthly_depr = (asset.cost - asset.salvage_value) / (asset.useful_life * 12)
        monthly_depr = round(monthly_depr, 2)
        
        max_depr_allowed = asset.cost - asset.salvage_value - asset.accumulated_depreciation
        if max_depr_allowed <= 0:
            continue
            
        depr_to_apply = min(monthly_depr, max_depr_allowed)
        if depr_to_apply > 0:
            asset.accumulated_depreciation += depr_to_apply
            total_depr_amount += depr_to_apply
            
    if total_depr_amount > 0:
        total_depr_amount = round(total_depr_amount, 2)
        
        # Post double entry: Dr Depreciation Expense / Cr Accumulated Depreciation
        je = db.JournalEntry(
            user_id=uid, date=period + "-28",
            description=f"Monthly Depreciation Expense - {period}",
            reference=f"DEPR-{period}", entry_type="adjustment"
        )
        session.add(je)
        session.flush()
        
        if depr_expense_account:
            session.add(db.JournalLine(entry_id=je.id, account_id=depr_expense_account.id, debit=total_depr_amount))
            depr_expense_account.balance += total_depr_amount
        if accum_depr_account:
            session.add(db.JournalLine(entry_id=je.id, account_id=accum_depr_account.id, credit=total_depr_amount))
            accum_depr_account.balance -= total_depr_amount
            
        session.commit()
        return JSONResponse({"success": True, "message": f"Depreciation executed: ${total_depr_amount:.2f} posted to adjustments."})
        
    session.commit()
    return JSONResponse({"success": True, "message": "No depreciation applied (all assets fully depreciated)."})


# ──────────────────────────────────────────────────────
# AI AGENT (Nexus AI)
# ──────────────────────────────────────────────────────
@app.get("/ai", response_class=HTMLResponse)
async def ai_page(request: Request):
    user = require_user(request)
    return templates.TemplateResponse(request, "ai_chat.html", {
        "user": user, "active_page": "ai"
    })

@app.post("/api/ai/chat")
async def ai_chat(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    uid = user["id"]
    body = await request.json()
    question = body.get("message", "").strip()
    
    if not question:
        return JSONResponse({"reply": "Please type a question."})
    
    # Build financial context from real data
    accounts = session.query(db.Account).filter(db.Account.user_id == uid).all()
    total_rev  = sum(a.balance for a in accounts if a.account_type == "Revenue")
    total_exp  = sum(a.balance for a in accounts if a.account_type == "Expense")
    net_profit = total_rev - total_exp
    cash       = next((a.balance for a in accounts if a.code == "1000"), 0)
    ar         = next((a.balance for a in accounts if a.code == "1200"), 0)
    
    context = f"""
    You are Nexus AI, an expert accounting and financial advisor integrated into the Nexus ERP system.
    
    Current Financial Summary for {user.get('company', 'the business')}:
    - Total Revenue: {user.get('currency','$')}{total_rev:,.2f}
    - Total Expenses: {user.get('currency','$')}{total_exp:,.2f}
    - Net Profit / (Loss): {user.get('currency','$')}{net_profit:,.2f}
    - Cash at Bank: {user.get('currency','$')}{cash:,.2f}
    - Accounts Receivable: {user.get('currency','$')}{ar:,.2f}
    - Business Plan: {user.get('plan', 'starter')}
    
    Respond in a helpful, professional, and easy-to-understand tone.
    If asked about accounting concepts, explain them clearly. 
    If asked for suggestions, base them on the actual financial data provided.
    Keep answers concise (2-4 sentences) unless a detailed explanation is required.
    """
    
    api_key = os.getenv("GEMINI_API_KEY", "")
    
    if api_key and api_key != "your_gemini_api_key_here":
        try:
            import urllib.request
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}"
            payload = json.dumps({
                "contents": [{"parts": [{"text": context + "\n\nUser Question: " + question}]}]
            }).encode()
            req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"})
            with urllib.request.urlopen(req, timeout=15) as response:
                result = json.loads(response.read())
                reply = result["candidates"][0]["content"]["parts"][0]["text"]
        except Exception as e:
            reply = _fallback_ai(question, total_rev, total_exp, net_profit, cash, ar)
    else:
        reply = _fallback_ai(question, total_rev, total_exp, net_profit, cash, ar)
    
    return JSONResponse({"reply": reply})

def _fallback_ai(question: str, revenue: float, expenses: float, profit: float, cash: float, ar: float) -> str:
    """Built-in rule-based AI responses when Gemini API is not configured."""
    q = question.lower()
    
    if any(w in q for w in ["profit", "loss", "income"]):
        status = "profitable" if profit >= 0 else "running at a loss"
        return f"Based on your current data, your business is {status} with a net {'profit' if profit >= 0 else 'loss'} of ${abs(profit):,.2f}. Revenue stands at ${revenue:,.2f} against expenses of ${expenses:,.2f}."
    
    elif any(w in q for w in ["cash", "liquidity", "bank"]):
        return f"Your current cash at bank balance is ${cash:,.2f}. {'This looks healthy!' if cash > 5000 else 'Consider reviewing your cash flow — your balance is relatively low.'}"
    
    elif any(w in q for w in ["receivable", "collection", "customer"]):
        return f"You have ${ar:,.2f} in outstanding Accounts Receivable. Follow up on overdue invoices to improve your cash position."
    
    elif any(w in q for w in ["depreciation", "what is", "explain"]):
        return "Depreciation is the systematic reduction in the book value of a fixed asset over its useful life. For example, if you buy a computer for $1,000 and it lasts 5 years, you expense $200/year. This matches expenses to the revenue they help generate — a core principle in accounting."
    
    elif any(w in q for w in ["improve", "suggestion", "advice", "better"]):
        tips = []
        if profit < 0: tips.append("📉 Your expenses exceed revenue — review your largest expense categories immediately.")
        if ar > cash * 2: tips.append("📬 You have high receivables vs cash — prioritize collecting from customers.")
        if cash < 1000: tips.append("💰 Your cash balance is critically low — consider a credit line or reduce outflows.")
        if not tips: tips.append("✅ Your finances look healthy! Continue monitoring your cash flow weekly.")
        return "\n".join(tips)
    
    elif any(w in q for w in ["balance sheet", "asset", "liability", "equity"]):
        return "A Balance Sheet follows the fundamental equation: Assets = Liabilities + Equity. It shows what your business OWNS (assets), what it OWES (liabilities), and the owner's net stake (equity) at a specific point in time."
    
    elif any(w in q for w in ["double entry", "journal", "debit", "credit"]):
        return "Double-entry accounting means every transaction affects at least two accounts — one is debited and one is credited by equal amounts. Rule: Assets and Expenses increase with Debits; Liabilities, Equity, and Revenue increase with Credits."
    
    else:
        return f"Great question! Based on your financials — Revenue: ${revenue:,.2f}, Expenses: ${expenses:,.2f}, Net Profit: ${profit:,.2f} — I'm here to help with any accounting question. Could you be more specific about what you'd like to know?"


# ──────────────────────────────────────────────────────
# CEO ADMIN PANEL & MANAGEMENT
# ──────────────────────────────────────────────────────
@app.get("/ceo", response_class=HTMLResponse)
async def ceo_dashboard(request: Request, session: Session = Depends(db.get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/auth", status_code=303)
    
    uid = user["id"]
    accounts = session.query(db.Account).filter(db.Account.user_id == uid).all()
    if not accounts:
        accounts = session.query(db.Account).all()

    total_assets      = sum(a.balance for a in accounts if a.account_type == "Asset" and a.balance > 0)
    total_liabilities = sum(a.balance for a in accounts if a.account_type == "Liability")
    total_revenue     = sum(a.balance for a in accounts if a.account_type == "Revenue")
    total_expenses    = sum(a.balance for a in accounts if a.account_type == "Expense")
    net_profit        = total_revenue - total_expenses
    ebitda_estimate   = net_profit + sum(a.balance for a in accounts if "Depreciation" in a.name or "Interest" in a.name)
    cash_reserve      = sum(a.balance for a in accounts if "Cash" in a.name)
    operating_burn    = total_expenses / 6.0 if total_expenses > 0 else 14500.0

    # Approvals queue (pending proposals)
    pending_approvals = session.query(db.CommunicationMessage).filter(
        db.CommunicationMessage.category == "Financial Proposal",
        db.CommunicationMessage.status == "Pending"
    ).order_by(db.CommunicationMessage.created_at.desc()).all()

    # Executive thoughts feed
    thoughts = session.query(db.CommunicationMessage).order_by(db.CommunicationMessage.created_at.desc()).limit(15).all()

    # User & Role management
    team_members = session.query(db.User).all()

    return templates.TemplateResponse(request, "ceo.html", {
        "user": user,
        "active_page": "ceo",
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "total_revenue": total_revenue,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "ebitda_estimate": ebitda_estimate,
        "cash_reserve": cash_reserve,
        "operating_burn": operating_burn,
        "pending_approvals": pending_approvals,
        "thoughts": thoughts,
        "team_members": team_members
    })

@app.post("/ceo/approvals")
async def action_approval(
    request: Request,
    message_id: int = Form(...),
    action: str = Form(...),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    msg = session.query(db.CommunicationMessage).filter(db.CommunicationMessage.id == message_id).first()
    if msg:
        msg.status = action
        session.commit()
    return RedirectResponse(url="/ceo", status_code=303)

@app.post("/ceo/users/role")
async def update_user_role(
    request: Request,
    target_user_id: int = Form(...),
    new_role: str = Form(...),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    target = session.query(db.User).filter(db.User.id == target_user_id).first()
    if target and new_role in ["CEO", "CFO", "Accountant"]:
        target.role = new_role
        session.commit()
    return RedirectResponse(url="/ceo", status_code=303)


# ──────────────────────────────────────────────────────
# CFO ADMIN PANEL & FINANCIAL CONTROLS
# ──────────────────────────────────────────────────────
@app.get("/cfo", response_class=HTMLResponse)
async def cfo_dashboard(request: Request, session: Session = Depends(db.get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/auth", status_code=303)

    uid = user["id"]
    accounts = session.query(db.Account).filter(db.Account.user_id == uid).all()
    if not accounts:
        accounts = session.query(db.Account).all()

    current_assets = sum(a.balance for a in accounts if a.account_type == "Asset" and a.sub_type == "Current Asset")
    current_liabilities = sum(a.balance for a in accounts if a.account_type == "Liability" and a.sub_type == "Current Liability")
    inventory = sum(a.balance for a in accounts if "Inventory" in a.name)
    cash_and_equiv = sum(a.balance for a in accounts if "Cash" in a.name)

    working_capital = current_assets - current_liabilities
    current_ratio = round(current_assets / current_liabilities, 2) if current_liabilities > 0 else 3.25
    quick_ratio = round((current_assets - inventory) / current_liabilities, 2) if current_liabilities > 0 else 2.80

    total_liabilities = sum(a.balance for a in accounts if a.account_type == "Liability")
    total_equity = sum(a.balance for a in accounts if a.account_type == "Equity")
    debt_to_equity = round(total_liabilities / total_equity, 2) if total_equity > 0 else 0.45

    my_proposals = session.query(db.CommunicationMessage).filter(
        db.CommunicationMessage.sender_role == "CFO"
    ).order_by(db.CommunicationMessage.created_at.desc()).all()

    unreconciled_count = session.query(db.BankTransaction).filter(
        db.BankTransaction.status == "unmatched"
    ).count()

    recent_logs = session.query(db.AuditLog).order_by(db.AuditLog.created_at.desc()).limit(8).all()

    return templates.TemplateResponse(request, "cfo.html", {
        "user": user,
        "active_page": "cfo",
        "current_assets": current_assets,
        "current_liabilities": current_liabilities,
        "working_capital": working_capital,
        "current_ratio": current_ratio,
        "quick_ratio": quick_ratio,
        "debt_to_equity": debt_to_equity,
        "cash_and_equiv": cash_and_equiv,
        "my_proposals": my_proposals,
        "unreconciled_count": unreconciled_count,
        "recent_logs": recent_logs
    })

@app.post("/cfo/proposals")
async def submit_cfo_proposal(
    request: Request,
    subject: str = Form(...),
    message: str = Form(...),
    amount: float = Form(0.0),
    urgency: str = Form("Important"),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    proposal = db.CommunicationMessage(
        user_id=user["id"],
        company_name=user.get("company", "Nexus Enterprise"),
        sender_username=user["username"],
        sender_role="CFO",
        recipient_role="CEO",
        subject=subject,
        message=message,
        category="Financial Proposal",
        urgency=urgency,
        status="Pending",
        amount=amount
    )
    session.add(proposal)
    session.commit()
    return RedirectResponse(url="/cfo", status_code=303)


# ──────────────────────────────────────────────────────
# TRI-PARTY EXECUTIVE COMMUNICATION SYSTEM
# ──────────────────────────────────────────────────────
@app.get("/communication", response_class=HTMLResponse)
async def communication_page(request: Request, session: Session = Depends(db.get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/auth", status_code=303)

    messages = session.query(db.CommunicationMessage).order_by(db.CommunicationMessage.created_at.desc()).all()
    return templates.TemplateResponse(request, "communication.html", {
        "user": user,
        "active_page": "communication",
        "messages": messages
    })

@app.post("/api/communication/send")
async def send_communication(
    request: Request,
    subject: str = Form(...),
    message: str = Form(...),
    recipient_role: str = Form("ALL"),
    category: str = Form("General Exchange"),
    urgency: str = Form("Normal"),
    amount: float = Form(0.0),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    sender_role = user.get("role", "Accountant")

    new_msg = db.CommunicationMessage(
        user_id=user["id"],
        company_name=user.get("company", "Nexus Enterprise"),
        sender_username=user["username"],
        sender_role=sender_role,
        recipient_role=recipient_role,
        subject=subject,
        message=message,
        category=category,
        urgency=urgency,
        status="Pending" if category == "Financial Proposal" else "Acknowledged",
        amount=amount
    )
    session.add(new_msg)
    session.commit()

    if request.headers.get("accept") == "application/json":
        return JSONResponse({"status": "success", "id": new_msg.id})
    return RedirectResponse(url=request.headers.get("referer", "/communication"), status_code=303)

@app.post("/api/communication/action")
async def action_message(
    request: Request,
    message_id: int = Form(...),
    action: str = Form(...),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    msg = session.query(db.CommunicationMessage).filter(db.CommunicationMessage.id == message_id).first()
    if msg:
        msg.status = action
        session.commit()
        return JSONResponse({"status": "success", "message_id": message_id, "new_status": action})
    return JSONResponse({"status": "error", "message": "Message not found"}, status_code=404)

@app.get("/api/communication/feed")
async def communication_feed(session: Session = Depends(db.get_db)):
    messages = session.query(db.CommunicationMessage).order_by(db.CommunicationMessage.created_at.desc()).limit(30).all()
    out = []
    for m in messages:
        out.append({
            "id": m.id,
            "sender_username": m.sender_username,
            "sender_role": m.sender_role,
            "recipient_role": m.recipient_role,
            "subject": m.subject,
            "message": m.message,
            "category": m.category,
            "urgency": m.urgency,
            "status": m.status,
            "amount": m.amount,
            "created_at": m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else ""
        })
    return JSONResponse({"messages": out})


# ──────────────────────────────────────────────────────
# SAAS BILLING, PAYMENT GATEWAY & PREFERENCES
# ──────────────────────────────────────────────────────
@app.get("/subscription/checkout", response_class=HTMLResponse)
async def checkout_page(request: Request, session: Session = Depends(db.get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/auth", status_code=303)

    db_user = session.query(db.User).filter(db.User.id == user["id"]).first()
    payments = session.query(db.PaymentTransaction).filter(
        db.PaymentTransaction.user_id == user["id"]
    ).order_by(db.PaymentTransaction.created_at.desc()).all()

    return templates.TemplateResponse(request, "checkout.html", {
        "user": user,
        "db_user": db_user,
        "active_page": "checkout",
        "payments": payments
    })

@app.post("/api/payment/process")
async def process_payment(
    request: Request,
    plan_name: str = Form(...),
    amount: float = Form(...),
    gateway: str = Form("Stripe"),
    card_name: str = Form(...),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    tx_ref = f"NEX-{uuid.uuid4().hex[:8].upper()}"

    # Record Payment Transaction
    payment = db.PaymentTransaction(
        user_id=user["id"],
        amount=amount,
        currency=user.get("currency", "USD"),
        plan_name=plan_name,
        gateway=gateway,
        status="success",
        transaction_ref=tx_ref
    )
    session.add(payment)

    # Upgrade User Subscription
    db_user = session.query(db.User).filter(db.User.id == user["id"]).first()
    if db_user:
        db_user.plan = plan_name
        db_user.subscription_status = "active"
        session.commit()

    # Update session user dict
    user["plan"] = plan_name
    user["subscription_status"] = "active"

    resp = RedirectResponse(url="/subscription/checkout?success=true", status_code=303)
    encoded = urllib.parse.quote(json.dumps(user))
    resp.set_cookie("nexus_user", encoded, max_age=86400 * 7)
    return resp

@app.post("/api/user/preferences")
async def update_user_preferences(
    request: Request,
    currency: str = Form(...),
    language: str = Form(...),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    symbol = db.CURRENCY_SYMBOLS.get(currency, "$")

    db_user = session.query(db.User).filter(db.User.id == user["id"]).first()
    if db_user:
        db_user.currency = currency
        db_user.currency_symbol = symbol
        db_user.language = language
        session.commit()

    user["currency"] = currency
    user["currency_symbol"] = symbol
    user["language"] = language

    resp = RedirectResponse(url=request.headers.get("referer", "/"), status_code=303)
    encoded = urllib.parse.quote(json.dumps(user))
    resp.set_cookie("nexus_user", encoded, max_age=86400 * 7)
    return resp


# ──────────────────────────────────────────────────────
# HR & CUSTOM TEAM SEATS (Partnership & Corporate Plans)
# ──────────────────────────────────────────────────────
@app.get("/team", response_class=HTMLResponse)
async def team_page(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    plan = user.get("plan", "starter")
    if plan == "starter":
        return RedirectResponse(url="/subscription/checkout?upgrade=required", status_code=303)

    members = session.query(db.TeamMember).filter(db.TeamMember.owner_id == user["id"]).all()
    return templates.TemplateResponse(request, "team.html", {
        "user": user,
        "members": members,
        "active_page": "team"
    })

@app.post("/api/team/invite")
async def invite_team_member(
    request: Request,
    member_name: str = Form(...),
    email: str = Form(...),
    role_title: str = Form("Accountant"),
    permissions: str = Form("read_write"),
    session: Session = Depends(db.get_db)
):
    user = require_user(request)
    new_member = db.TeamMember(
        owner_id=user["id"],
        member_name=member_name,
        email=email,
        role_title=role_title,
        permissions=permissions
    )
    session.add(new_member)
    session.commit()
    return RedirectResponse(url="/team?success=true", status_code=303)


# ──────────────────────────────────────────────────────
# APP CREATOR SUPER-ADMIN MASTER PANEL (Private Ledger Shield)
# ──────────────────────────────────────────────────────
@app.get("/superadmin", response_class=HTMLResponse)
async def superadmin_panel(request: Request, session: Session = Depends(db.get_db)):
    user = require_user(request)
    if not user.get("is_superadmin"):
        raise HTTPException(status_code=403, detail="Access denied. App Creator SuperAdmin authorization required.")

    all_users = session.query(db.User).order_by(db.User.created_at.desc()).all()
    all_payments = session.query(db.PaymentTransaction).order_by(db.PaymentTransaction.created_at.desc()).all()
    total_revenue = sum(p.amount for p in all_payments if p.status == "success")

    return templates.TemplateResponse(request, "superadmin.html", {
        "user": user,
        "all_users": all_users,
        "all_payments": all_payments,
        "total_revenue": total_revenue,
        "total_user_count": len(all_users),
        "active_page": "superadmin"
    })



